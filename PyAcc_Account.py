from PyAcc_Transactions import *
from openpyxl import Workbook

transaction = Transaction()
wb = Workbook()
# Use AcctClassses
# worksheets = accounts
# edit create worksheet to setup frrame work of blanace tracking

#########
# Python Accountant Class Definitions
# Desc: Contains Class definitions used in PyAcc.py and AccountantMain.py
# Created By: Samuel Buzas & Elizabeth Sheetz
#########




class Account:

	def __init__(self, account_Name= '(Chose an Account)' , workbook_name = '(Chose A Wokbook)', initial_balance = 0):
		self.workbook_name = str(workbook_name) + '.xlsx'
		self.account_Name = str(account_Name)
		self.initial_balance = initial_balance

	def balance_update(self):
		wb = load_workbook(self.workbook_name)
		ws = wb[self.account_Name]
		ws['G2'] = "=SUM(C2: C" + str(transaction.checkEntries(self.workbook_name,self.account_Name)) +")"+"+" + str(self.initial_balance)
		wb.save(self.workbook_name)

	# def get_transactions(self):
	# 	#print list of transaction numbers
	# 	print(t.entryNumber for t in self.transactions)

	# def add_transaction(self, tnew):
	# 	#add transaction to transaction list and update balance
	# 	self.transactions += [tnew]
	# 	self.balance_update(tnew.amount)

	def create_new_Account(self):
		#Use to create and populate a new account structure
		wb = load_workbook(self.workbook_name)
		wb.create_sheet(title = self.account_Name)
		sheet = wb[self.account_Name]
		sheet['A1'] = 'Date'
		sheet.column_dimensions['A'].width = 17
		sheet['B1'] = 'Category'
		sheet.column_dimensions['B'].width = 15
		sheet['C1'] = 'Amount'
		sheet.column_dimensions['C'].width = 10
		sheet['D1'] = 'Check Number'
		sheet.column_dimensions['D'].width = 17
		sheet['E1'] = 'Description'
		sheet.column_dimensions['E'].width = 50
		wb.geuss_types = True
		sheet['G1'] = self.account_Name + ' Balance'
		sheet.column_dimensions['G'].width = 25
		wb.save(self.workbook_name)
		sheet['G2'] = "=SUM(C2:C" + str(transaction.checkEntries(self.workbook_name,self.account_Name)) +")" + "+" + str(self.initial_balance)
		wb.save(self.workbook_name)

